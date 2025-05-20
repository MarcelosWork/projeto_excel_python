import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"

# Estilo de bordas finas pretas para a tabela TOTAIS
thin_side = Side(style="thin", color="000000")
border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)


def substituir_placeholders(sheet, produto):
    """
    Substitui placeholders do produto (e apenas do produto) na folha.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for chave, valor in produto.items():
                    placeholder = f"{{{{{chave}}}}}"
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(valor or ""))


def gerar_workbook(produtos, obra, cliente):
    # 1. Abre o template
    wb = load_workbook(TEMPLATE_PATH, data_only=False)

    # 2. Isola a folha V1 como modelo temporário
    tpl = wb["V1"]
    tpl.title = "_template_"

    # 3. Copia e preenche V1…VN para cada produto
    for idx, prod in enumerate(produtos):
        nova = wb.copy_worksheet(wb["_template_"])
        nova.title = f"V{idx+1}"
        substituir_placeholders(nova, prod)

    # 4. Remove a folha-templante
    wb.remove(wb["_template_"])

    # 5. Substitui {{obra}} e {{cliente}} em todas as folhas, incluindo TOTAIS
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("{{obra}}", obra)
                    cell.value = cell.value.replace("{{cliente}}", cliente)

    # 6. Preenche a tabela na folha TOTAIS
    totais = wb["TOTAIS"]
    header_row = 9
    start_row = header_row + 1

    # limpa linhas antigas
    if totais.max_row >= start_row:
        totais.delete_rows(start_row, totais.max_row - start_row + 1)

    # insere uma linha por produto
    for idx, prod in enumerate(produtos):
        r = start_row + idx
        totais.insert_rows(r)
        sheet_name = f"V{idx+1}"
        totais.cell(r, 1).value = prod.get("ref_janela")
        totais.cell(r, 2).value = prod.get("descricao")
        totais.cell(r, 3).value = prod.get("quantidade")
        totais.cell(r, 4).value = prod.get("altura")
        totais.cell(r, 5).value = prod.get("largura")
        totais.cell(r, 6).value  = f"='{sheet_name}'!J12"
        totais.cell(r, 7).value  = f"='{sheet_name}'!J12"
        totais.cell(r, 8).value  = f"='{sheet_name}'!J13"
        totais.cell(r, 9).value  = f"='{sheet_name}'!J13"
        totais.cell(r,10).value = prod.get("mao_obra_producao")
        totais.cell(r,11).value = prod.get("mao_obra_montagem")
        totais.cell(r,12).value = f"='{sheet_name}'!J35"
        totais.cell(r,13).value = f"='{sheet_name}'!J35"
        totais.cell(r,14).value = f"='{sheet_name}'!D35"
        # formatar números e euro
        totais.cell(r,4).number_format  = "0.000"
        totais.cell(r,5).number_format  = "0.000"
        for c in (6,7,8,9,12,13,14):
            totais.cell(r,c).number_format = "0.00 €"
        totais.cell(r,10).number_format = "0.00"
        totais.cell(r,11).number_format = "0.00"
        # aplicar bordas
        for c in range(1,15):
            totais.cell(r,c).border = border

    # 7. Inserir linha de totais com espaçamento
    end_row = start_row + len(produtos) - 1
    tot_row  = end_row + 2
    totais.insert_rows(tot_row)
    totais.cell(tot_row,1).value = "TOTAIS"
    totais.cell(tot_row,7).value  = f"=SUM(G{start_row}:G{end_row})"
    totais.cell(tot_row,9).value  = f"=SUM(I{start_row}:I{end_row})"
    totais.cell(tot_row,10).value = f"=SUM(J{start_row}:J{end_row})"
    totais.cell(tot_row,11).value = f"=SUM(K{start_row}:K{end_row})"
    totais.cell(tot_row,13).value = f"=SUM(M{start_row}:M{end_row})"
    # aplicar formatos e bordas aos totais
    for c, fmt in [(7,"0.00 €"),(9,"0.00 €"),(10,"0.00"),(11,"0.00"),(13,"0.00 €")]:
        totais.cell(tot_row,c).number_format = fmt
        totais.cell(tot_row,c).border = border

    return wb

